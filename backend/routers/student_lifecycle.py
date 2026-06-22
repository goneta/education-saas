from __future__ import annotations

import csv
import io
import json
import secrets
from datetime import datetime, timezone
from xml.etree import ElementTree

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from .. import audit, database, models, rbac, schemas, security
from ..services import school_context, student_lifecycle


router = APIRouter(prefix="/student-lifecycle", tags=["Student lifecycle"])


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _profile_payload(db: Session, profile: models.StudentGlobalProfile, current_user: models.User) -> dict:
    enrollments = db.query(models.StudentEnrollment).options(
        joinedload(models.StudentEnrollment.school),
        joinedload(models.StudentEnrollment.school_model_assignment).joinedload(models.SchoolModelAssignment.school_model),
        joinedload(models.StudentEnrollment.academic_year),
        joinedload(models.StudentEnrollment.class_),
        joinedload(models.StudentEnrollment.program),
    ).filter(
        models.StudentEnrollment.student_global_profile_id == profile.id
    ).order_by(models.StudentEnrollment.start_date.desc(), models.StudentEnrollment.id.desc()).all()
    visible = []
    for enrollment in enrollments:
        owns_context = (
            current_user.role == models.UserRole.SUPER_ADMIN
            or school_context.user_can_access_school(db, current_user, enrollment.school)
        )
        transferred_access = db.query(models.StudentTransferRequest.id).filter(
            models.StudentTransferRequest.student_global_profile_id == profile.id,
            models.StudentTransferRequest.to_school_id == current_user.school_id,
            models.StudentTransferRequest.status.in_(["approved", "completed"]),
        ).first()
        if not owns_context and not transferred_access and profile.user_id != current_user.id:
            continue
        grades = db.query(models.Grade).options(
            joinedload(models.Grade.assessment).joinedload(models.Assessment.subject)
        ).filter(
            models.Grade.student_enrollment_id == enrollment.id
        ).order_by(models.Grade.id).all()
        attendance = db.query(models.Attendance).filter(
            models.Attendance.student_enrollment_id == enrollment.id
        ).order_by(models.Attendance.date.desc()).all()
        internships = db.query(models.InternshipAssignment).options(
            joinedload(models.InternshipAssignment.internship)
        ).filter(
            models.InternshipAssignment.student_enrollment_id == enrollment.id
        ).order_by(models.InternshipAssignment.id.desc()).all()
        certificates = db.query(models.CertificateRequest).filter(
            models.CertificateRequest.student_enrollment_id == enrollment.id
        ).order_by(models.CertificateRequest.generated_at.desc()).all()
        documents = db.query(models.StudentRegistrationDocument).filter(
            models.StudentRegistrationDocument.student_enrollment_id == enrollment.id
        ).order_by(models.StudentRegistrationDocument.id).all()
        financial = None
        if owns_context:
            invoices = db.query(models.StudentInvoice).filter(
                models.StudentInvoice.student_enrollment_id == enrollment.id,
                models.StudentInvoice.school_id == enrollment.school_id,
            ).all()
            financial = {
                "invoice_count": len(invoices),
                "amount_due": sum(row.amount_due for row in invoices),
                "amount_paid": sum(row.amount_paid for row in invoices),
                "remaining_balance": sum(row.remaining_balance for row in invoices),
            }
        lock = db.query(models.AcademicYearLock).filter(
            models.AcademicYearLock.school_id == enrollment.school_id,
            models.AcademicYearLock.academic_year_id == enrollment.academic_year_id,
        ).order_by(models.AcademicYearLock.id.desc()).first()
        visible.append({
            "id": enrollment.id,
            "organization_id": enrollment.organization_id,
            "school_id": enrollment.school_id,
            "school_name": enrollment.school.name,
            "school_model_assignment_id": enrollment.school_model_assignment_id,
            "model_code": enrollment.school_model_assignment.school_model.code,
            "model_name": enrollment.school_model_assignment.display_name or enrollment.school_model_assignment.school_model.name,
            "academic_year_id": enrollment.academic_year_id,
            "academic_year_name": enrollment.academic_year.name,
            "class_id": enrollment.class_id,
            "class_name": enrollment.class_.name if enrollment.class_ else None,
            "program_id": enrollment.program_id,
            "program_name": enrollment.program.name if enrollment.program else None,
            "enrollment_status": enrollment.enrollment_status,
            "enrollment_type": enrollment.enrollment_type,
            "schedule_type": enrollment.schedule_type,
            "primary_enrollment": enrollment.primary_enrollment,
            "allows_concurrent_enrollment": enrollment.allows_concurrent_enrollment,
            "days_of_week": enrollment.days_of_week or [],
            "start_time": enrollment.start_time.isoformat() if enrollment.start_time else None,
            "end_time": enrollment.end_time.isoformat() if enrollment.end_time else None,
            "start_date": enrollment.start_date.isoformat() if enrollment.start_date else None,
            "end_date": enrollment.end_date.isoformat() if enrollment.end_date else None,
            "location": enrollment.location,
            "academic_summary": {
                "grades": len(grades),
                "attendance_records": len(attendance),
                "internships": len(internships),
            },
            "academic_data": {
                "grades": [
                    {
                        "assessment": grade.assessment.title,
                        "subject": grade.assessment.subject.name if grade.assessment.subject else None,
                        "score": grade.score,
                        "maximum": grade.assessment.max_score,
                        "comment": grade.comment,
                    }
                    for grade in grades
                ],
                "attendance": [
                    {
                        "date": row.date,
                        "status": row.status.value if hasattr(row.status, "value") else row.status,
                        "remarks": row.remarks,
                    }
                    for row in attendance
                ],
                "internships": [
                    {
                        "title": row.internship.title,
                        "company": row.internship.company_name,
                        "status": row.status,
                    }
                    for row in internships
                ],
                "certificates": [
                    {
                        "type": row.certificate_type.value if hasattr(row.certificate_type, "value") else row.certificate_type,
                        "status": row.status.value if hasattr(row.status, "value") else row.status,
                        "generated_at": row.generated_at,
                    }
                    for row in certificates
                ],
                "documents": [
                    {"name": row.name, "is_received": row.is_received, "received_at": row.received_at}
                    for row in documents
                ],
            },
            "financial": financial,
            "read_only": bool(lock and lock.status in {"closed", "archived"}) or not owns_context,
            "lock_status": lock.status if lock else "open",
        })
    transfers = db.query(models.StudentTransferRequest).filter(
        models.StudentTransferRequest.student_global_profile_id == profile.id
    ).order_by(models.StudentTransferRequest.created_at.desc()).all()
    return {
        "id": profile.id,
        "user_id": profile.user_id,
        "student_profile_id": profile.student_profile_id,
        "global_student_number": profile.global_student_number,
        "first_name": profile.first_name,
        "last_name": profile.last_name,
        "full_name": f"{profile.first_name} {profile.last_name}".strip(),
        "date_of_birth": profile.date_of_birth,
        "gender": profile.gender,
        "nationality": profile.nationality,
        "photo_url": profile.photo_url,
        "enrollments": visible,
        "transfers": [
            {
                "id": row.id,
                "from_school_id": row.from_school_id,
                "to_school_id": row.to_school_id,
                "status": row.status,
                "academic_data_access_level": row.academic_data_access_level,
                "financial_data_access_allowed": False,
                "created_at": row.created_at,
                "approved_at": row.approved_at,
                "completed_at": row.completed_at,
            }
            for row in transfers
            if current_user.role == models.UserRole.SUPER_ADMIN
            or current_user.school_id in {row.from_school_id, row.to_school_id}
            or profile.user_id == current_user.id
        ],
    }


@router.get("/search")
def search_global_students(
    q: str = Query(min_length=2),
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:view", db)
    pattern = f"%{q.strip()}%"
    rows = db.query(models.StudentGlobalProfile).join(models.StudentProfile).join(models.User).filter(
        or_(
            models.StudentGlobalProfile.global_student_number.ilike(pattern),
            models.StudentGlobalProfile.first_name.ilike(pattern),
            models.StudentGlobalProfile.last_name.ilike(pattern),
            models.StudentProfile.registration_number.ilike(pattern),
            models.StudentProfile.parent_name.ilike(pattern),
        )
    ).order_by(models.StudentGlobalProfile.last_name, models.StudentGlobalProfile.first_name).limit(30).all()
    return [
        {
            "id": row.id,
            "global_student_number": row.global_student_number,
            "full_name": f"{row.first_name} {row.last_name}".strip(),
            "date_of_birth": row.date_of_birth,
            "parent_name": row.student_profile.parent_name,
        }
        for row in rows
    ]


@router.get("/students/{student_user_id}")
def global_student_profile(
    student_user_id: int,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    student = db.query(models.StudentProfile).filter(
        models.StudentProfile.user_id == student_user_id
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="Eleve introuvable.")
    profile = student_lifecycle.ensure_global_profile(db, student)
    student_lifecycle.ensure_student_context_access(
        db,
        current_user=current_user,
        global_profile_id=profile.id,
    )
    db.commit()
    return _profile_payload(db, profile, current_user)


@router.post("/enrollments")
def create_enrollment(
    payload: schemas.StudentEnrollmentCreate,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:create", db)
    context = school_context.resolve_context(
        db,
        current_user,
        school_model_assignment_id=payload.school_model_assignment_id,
        academic_year_id=payload.academic_year_id,
    )
    if not context.academic_year_id:
        year = student_lifecycle.ensure_academic_year_for_context(
            db,
            school_id=context.school_id,
            school_model_assignment_id=context.school_model_assignment_id,
        )
        context = school_context.resolve_context(
            db,
            current_user,
            school_model_assignment_id=context.school_model_assignment_id,
            academic_year_id=year.id,
        )
    profile = None
    if payload.student_global_profile_id:
        profile = db.query(models.StudentGlobalProfile).filter(
            models.StudentGlobalProfile.id == payload.student_global_profile_id
        ).first()
    elif payload.student_user_id:
        student = db.query(models.StudentProfile).filter(
            models.StudentProfile.user_id == payload.student_user_id
        ).first()
        if student:
            profile = student_lifecycle.ensure_global_profile(db, student)
    if not profile:
        raise HTTPException(status_code=404, detail="Profil global eleve introuvable.")
    assignment = db.query(models.SchoolModelAssignment).options(
        joinedload(models.SchoolModelAssignment.school_model),
        joinedload(models.SchoolModelAssignment.school),
    ).filter(models.SchoolModelAssignment.id == context.school_model_assignment_id).first()
    student_lifecycle.ensure_academic_year_is_editable(
        db,
        current_user=current_user,
        school_id=context.school_id,
        academic_year_id=context.academic_year_id,
        school_model_assignment_id=context.school_model_assignment_id,
        student_global_profile_id=profile.id,
        resource_type="student_enrollment",
    )
    conflicts = student_lifecycle.validate_concurrent_enrollment(
        db,
        global_profile_id=profile.id,
        assignment=assignment,
        academic_year_id=context.academic_year_id,
        enrollment_type=payload.enrollment_type,
        allows_concurrent_enrollment=payload.allows_concurrent_enrollment,
        days_of_week=payload.days_of_week,
        start_time=payload.start_time,
        end_time=payload.end_time,
        force=payload.force,
        override_reason=payload.override_reason,
        current_user=current_user,
    )
    year = db.query(models.AcademicYear).filter(models.AcademicYear.id == context.academic_year_id).first()
    row = models.StudentEnrollment(
        student_global_profile_id=profile.id,
        organization_id=context.organization_id,
        school_id=context.school_id,
        school_model_assignment_id=context.school_model_assignment_id,
        academic_year_id=context.academic_year_id,
        class_id=payload.class_id,
        level_id=payload.level_id,
        program_id=payload.program_id,
        enrollment_status="active",
        enrollment_type=payload.enrollment_type,
        schedule_type=payload.schedule_type,
        allows_concurrent_enrollment=payload.allows_concurrent_enrollment,
        primary_enrollment=payload.primary_enrollment,
        module_id=payload.module_id,
        training_program_id=payload.training_program_id,
        certification_id=payload.certification_id,
        start_date=payload.start_date or year.start_date or _now(),
        end_date=payload.end_date or year.end_date,
        start_time=payload.start_time,
        end_time=payload.end_time,
        days_of_week=payload.days_of_week,
        location=payload.location,
        created_by_user_id=current_user.id,
        override_reason=payload.override_reason,
    )
    db.add(row)
    db.flush()
    if payload.primary_enrollment:
        db.query(models.StudentEnrollment).filter(
            models.StudentEnrollment.student_global_profile_id == profile.id,
            models.StudentEnrollment.id != row.id,
        ).update({"primary_enrollment": False}, synchronize_session=False)
        profile.student_profile.current_class_id = payload.class_id
        profile.student_profile.school_model_assignment_id = context.school_model_assignment_id
    audit.record_audit(
        db,
        action="student_enrollment.created",
        current_user=current_user,
        entity_type="student_enrollment",
        entity_id=row.id,
        details={"global_profile_id": profile.id, "conflicts_overridden": conflicts},
    )
    db.commit()
    return {"id": row.id, "status": row.enrollment_status, "conflicting_enrollment_ids": conflicts}


@router.post("/transfers")
def request_transfer(
    payload: schemas.StudentTransferCreate,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:edit", db)
    source = db.query(models.StudentEnrollment).filter(
        models.StudentEnrollment.id == payload.from_enrollment_id,
        models.StudentEnrollment.student_global_profile_id == payload.student_global_profile_id,
    ).first()
    destination_assignment = db.query(models.SchoolModelAssignment).options(
        joinedload(models.SchoolModelAssignment.school)
    ).filter(models.SchoolModelAssignment.id == payload.to_school_model_assignment_id).first()
    destination_year = db.query(models.AcademicYear).filter(
        models.AcademicYear.id == payload.to_academic_year_id,
        models.AcademicYear.school_id == destination_assignment.school_id if destination_assignment else -1,
    ).first()
    if not source or not destination_assignment or not destination_year:
        raise HTTPException(status_code=404, detail="Contexte de transfert introuvable.")
    source_school = db.query(models.School).filter(models.School.id == source.school_id).first()
    if not (
        current_user.role == models.UserRole.SUPER_ADMIN
        or school_context.user_can_access_school(db, current_user, source_school)
        or school_context.user_can_access_school(db, current_user, destination_assignment.school)
    ):
        raise HTTPException(status_code=403, detail="Transfert hors de votre perimetre.")
    duplicate = db.query(models.StudentTransferRequest.id).filter(
        models.StudentTransferRequest.student_global_profile_id == payload.student_global_profile_id,
        models.StudentTransferRequest.from_school_id == source.school_id,
        models.StudentTransferRequest.to_school_id == destination_assignment.school_id,
        models.StudentTransferRequest.status == "pending",
    ).first()
    if duplicate:
        raise HTTPException(status_code=409, detail="Une demande de transfert est deja en attente.")
    row = models.StudentTransferRequest(
        student_global_profile_id=payload.student_global_profile_id,
        from_organization_id=source.organization_id,
        from_school_id=source.school_id,
        from_school_model_assignment_id=source.school_model_assignment_id,
        from_academic_year_id=source.academic_year_id,
        to_organization_id=destination_assignment.school.organization_id,
        to_school_id=destination_assignment.school_id,
        to_school_model_assignment_id=destination_assignment.id,
        to_academic_year_id=destination_year.id,
        requested_by_user_id=current_user.id,
        academic_data_access_level=payload.academic_data_access_level,
        financial_data_access_allowed=False,
        notes=payload.notes,
    )
    db.add(row)
    db.flush()
    audit.record_audit(
        db,
        action="student_transfer.requested",
        current_user=current_user,
        entity_type="student_transfer_request",
        entity_id=row.id,
        details={"from_school_id": row.from_school_id, "to_school_id": row.to_school_id},
    )
    db.commit()
    return {"id": row.id, "status": row.status}


@router.get("/transfers")
def list_transfers(
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:view", db)
    query = db.query(models.StudentTransferRequest)
    if current_user.role != models.UserRole.SUPER_ADMIN:
        query = query.filter(
            or_(
                models.StudentTransferRequest.from_school_id == current_user.school_id,
                models.StudentTransferRequest.to_school_id == current_user.school_id,
            )
        )
    return [
        {
            "id": row.id,
            "student_global_profile_id": row.student_global_profile_id,
            "global_student_number": row.student_global_profile.global_student_number,
            "student_name": f"{row.student_global_profile.first_name} {row.student_global_profile.last_name}".strip(),
            "from_school_id": row.from_school_id,
            "to_school_id": row.to_school_id,
            "status": row.status,
            "created_at": row.created_at,
        }
        for row in query.options(joinedload(models.StudentTransferRequest.student_global_profile)).order_by(
            models.StudentTransferRequest.created_at.desc()
        ).all()
    ]


@router.post("/transfers/{transfer_id}/decision")
def decide_transfer(
    transfer_id: int,
    payload: schemas.StudentTransferDecision,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:approve", db)
    row = db.query(models.StudentTransferRequest).filter(
        models.StudentTransferRequest.id == transfer_id
    ).first()
    if not row:
        raise HTTPException(status_code=404, detail="Transfert introuvable.")
    student_lifecycle.ensure_transfer_access(db, current_user=current_user, transfer=row)
    decision = payload.decision.lower()
    if decision not in {"approved", "rejected", "cancelled", "completed"}:
        raise HTTPException(status_code=400, detail="Decision invalide.")
    row.status = decision
    row.notes = payload.notes or row.notes
    if decision in {"approved", "completed"}:
        row.approved_by_user_id = current_user.id
        row.approved_at = _now()
    if decision == "completed":
        source = db.query(models.StudentEnrollment).filter(
            models.StudentEnrollment.student_global_profile_id == row.student_global_profile_id,
            models.StudentEnrollment.school_id == row.from_school_id,
            models.StudentEnrollment.academic_year_id == row.from_academic_year_id,
        ).order_by(models.StudentEnrollment.id.desc()).first()
        if source:
            source.enrollment_status = "transferred_out"
            source.transfer_to_school_id = row.to_school_id
            source.end_date = _now()
        existing = student_lifecycle.enrollment_for_context(
            db,
            row.student_global_profile_id,
            school_id=row.to_school_id,
            school_model_assignment_id=row.to_school_model_assignment_id,
            academic_year_id=row.to_academic_year_id,
        )
        if not existing:
            year = db.query(models.AcademicYear).filter(models.AcademicYear.id == row.to_academic_year_id).first()
            db.add(models.StudentEnrollment(
                student_global_profile_id=row.student_global_profile_id,
                organization_id=row.to_organization_id,
                school_id=row.to_school_id,
                school_model_assignment_id=row.to_school_model_assignment_id,
                academic_year_id=row.to_academic_year_id,
                class_id=payload.class_id,
                program_id=payload.program_id,
                enrollment_status="transferred_in",
                enrollment_type=payload.enrollment_type,
                schedule_type=payload.schedule_type,
                primary_enrollment=True,
                start_date=year.start_date or _now(),
                end_date=year.end_date,
                transfer_from_school_id=row.from_school_id,
                created_by_user_id=current_user.id,
            ))
        row.completed_at = _now()
    audit.record_audit(
        db,
        action=f"student_transfer.{decision}",
        current_user=current_user,
        entity_type="student_transfer_request",
        entity_id=row.id,
        details={"financial_data_access_allowed": False},
    )
    db.commit()
    return {"id": row.id, "status": row.status}


@router.post("/academic-years/{academic_year_id}/close")
def close_academic_year(
    academic_year_id: int,
    payload: schemas.AcademicYearCloseRequest,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "settings:manage_settings", db)
    if payload.confirmation.strip().upper() != "CLOTURER":
        raise HTTPException(status_code=400, detail="Saisissez CLOTURER pour confirmer.")
    context = school_context.resolve_context(
        db,
        current_user,
        school_model_assignment_id=payload.school_model_assignment_id,
        academic_year_id=academic_year_id,
    )
    row = db.query(models.AcademicYearLock).filter(
        models.AcademicYearLock.school_id == context.school_id,
        models.AcademicYearLock.school_model_assignment_id == context.school_model_assignment_id,
        models.AcademicYearLock.academic_year_id == academic_year_id,
    ).first()
    if not row:
        row = models.AcademicYearLock(
            organization_id=context.organization_id,
            school_id=context.school_id,
            school_model_assignment_id=context.school_model_assignment_id,
            academic_year_id=academic_year_id,
        )
        db.add(row)
    row.status = "closed"
    row.closed_at = _now()
    row.closed_by_user_id = current_user.id
    audit.record_audit(
        db,
        action="academic_year.closed",
        current_user=current_user,
        entity_type="academic_year",
        entity_id=academic_year_id,
        details={"confirmation": "CLOTURER"},
    )
    db.commit()
    return {"academic_year_id": academic_year_id, "status": "closed", "closed_at": row.closed_at}


@router.post("/historical-edit-grants")
def create_historical_edit_grant(
    payload: schemas.HistoricalEditGrantCreate,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    if current_user.role != models.UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Super Administrateur uniquement.")
    if payload.valid_until <= _now():
        raise HTTPException(status_code=400, detail="La date d'expiration doit etre future.")
    row = models.HistoricalDataEditGrant(
        granted_by_super_admin_id=current_user.id,
        valid_from=_now(),
        **payload.model_dump(),
    )
    db.add(row)
    db.flush()
    audit.record_audit(
        db,
        action="historical_edit_grant.created",
        current_user=current_user,
        entity_type="historical_data_edit_grant",
        entity_id=row.id,
        details={"reason": row.reason, "valid_until": row.valid_until.isoformat()},
    )
    db.commit()
    return {"id": row.id, "valid_until": row.valid_until, "is_active": row.is_active}


def _export_rows(payload: dict) -> list[dict]:
    return [
        {
            "global_student_number": payload["global_student_number"],
            "student_name": payload["full_name"],
            "academic_year": row["academic_year_name"],
            "school": row["school_name"],
            "school_model": row["model_name"],
            "class": row["class_name"] or "",
            "program": row["program_name"] or "",
            "status": row["enrollment_status"],
            "enrollment_type": row["enrollment_type"],
            "read_only": row["read_only"],
        }
        for row in payload["enrollments"]
    ]


@router.get("/students/{student_user_id}/export")
def export_student_journey(
    student_user_id: int,
    format: str = Query(default="pdf", pattern="^(pdf|csv|xlsx|markdown|xml|json)$"),
    include_finance: bool = False,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:export", db)
    student = db.query(models.StudentProfile).filter(models.StudentProfile.user_id == student_user_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Eleve introuvable.")
    profile = student_lifecycle.ensure_global_profile(db, student)
    student_lifecycle.ensure_student_context_access(
        db, current_user=current_user, global_profile_id=profile.id
    )
    payload = _profile_payload(db, profile, current_user)
    if not include_finance:
        for enrollment in payload["enrollments"]:
            enrollment["financial"] = None
    rows = _export_rows(payload)
    audit.record_audit(
        db,
        action="student_journey.exported",
        current_user=current_user,
        entity_type="student_global_profile",
        entity_id=profile.id,
        details={"format": format, "include_finance": include_finance},
    )
    db.commit()
    filename = f"parcours-{profile.global_student_number}.{format if format != 'markdown' else 'md'}"
    if format == "json":
        data = json.dumps(payload, default=str, ensure_ascii=False, indent=2).encode("utf-8")
        media_type = "application/json"
    elif format == "csv":
        stream = io.StringIO()
        writer = csv.DictWriter(stream, fieldnames=list(rows[0].keys()) if rows else ["global_student_number"])
        writer.writeheader()
        writer.writerows(rows)
        data = stream.getvalue().encode("utf-8-sig")
        media_type = "text/csv"
    elif format == "markdown":
        headers = list(rows[0].keys()) if rows else ["global_student_number"]
        lines = [f"# Parcours scolaire - {payload['full_name']}", "", "| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
        lines.extend("| " + " | ".join(str(row.get(key, "")) for key in headers) + " |" for row in rows)
        data = "\n".join(lines).encode("utf-8")
        media_type = "text/markdown"
    elif format == "xml":
        root = ElementTree.Element("studentJourney", exportedAt=_now().isoformat())
        identity = ElementTree.SubElement(root, "identity")
        for key in ("global_student_number", "full_name", "date_of_birth", "gender"):
            ElementTree.SubElement(identity, key).text = str(payload.get(key) or "")
        enrollments_node = ElementTree.SubElement(root, "enrollments")
        for row in rows:
            node = ElementTree.SubElement(enrollments_node, "enrollment")
            for key, value in row.items():
                ElementTree.SubElement(node, key).text = str(value)
        data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        media_type = "application/xml"
    elif format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Parcours scolaire"
        headers = list(rows[0].keys()) if rows else ["global_student_number"]
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(key) for key in headers])
        output = io.BytesIO()
        workbook.save(output)
        data = output.getvalue()
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        output = io.BytesIO()
        pdf = canvas.Canvas(output, pagesize=A4)
        school = db.query(models.School).filter(models.School.id == current_user.school_id).first()
        y = 810
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(40, y, school.name if school else "TeducAI")
        y -= 25
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y, f"Parcours scolaire - {payload['full_name']}")
        y -= 20
        pdf.setFont("Helvetica", 9)
        pdf.drawString(40, y, f"Numero global: {profile.global_student_number} | Exporte le {_now().strftime('%d/%m/%Y %H:%M')}")
        for row in rows:
            y -= 28
            if y < 60:
                pdf.showPage()
                y = 810
            pdf.drawString(40, y, f"{row['academic_year']} - {row['school']} - {row['school_model']}")
            pdf.drawString(55, y - 12, f"{row['class']} {row['program']} | {row['status']} | {row['enrollment_type']}")
        pdf.save()
        data = output.getvalue()
        media_type = "application/pdf"
    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_import_file(filename: str, content: bytes) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        values = list(workbook.active.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:]]
    if suffix == "json":
        payload = json.loads(content.decode("utf-8"))
        return payload if isinstance(payload, list) else payload.get("students", [])
    if suffix == "xml":
        root = ElementTree.fromstring(content)
        return [{child.tag: child.text for child in node} for node in root.findall(".//student")]
    if suffix in {"md", "markdown"}:
        lines = [line.strip() for line in content.decode("utf-8").splitlines() if line.strip().startswith("|")]
        if len(lines) < 3:
            return []
        headers = [value.strip() for value in lines[0].strip("|").split("|")]
        return [
            dict(zip(headers, [value.strip() for value in line.strip("|").split("|")]))
            for line in lines[2:]
        ]
    raise HTTPException(status_code=415, detail="Format accepte: CSV, XLSX, JSON, XML ou Markdown.")


@router.post("/imports/preview")
async def preview_student_import(
    file: UploadFile = File(...),
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:import", db)
    context = school_context.resolve_context(db, current_user)
    if not context.academic_year_id:
        year = student_lifecycle.ensure_academic_year_for_context(
            db,
            school_id=context.school_id,
            school_model_assignment_id=context.school_model_assignment_id,
        )
        context = school_context.resolve_context(
            db,
            current_user,
            school_model_assignment_id=context.school_model_assignment_id,
            academic_year_id=year.id,
        )
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Fichier trop volumineux (10 Mo maximum).")
    rows = _parse_import_file(file.filename or "import.csv", content)
    preview = []
    errors = []
    duplicate_count = 0
    for index, raw in enumerate(rows, start=2):
        row = {str(key).strip().lower(): value for key, value in raw.items()}
        full_name = str(row.get("full_name") or row.get("nom") or "").strip()
        registration = str(row.get("registration_number") or row.get("matricule") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        if not full_name or not registration:
            errors.append({"row": index, "message": "Nom complet et matricule obligatoires."})
            continue
        existing = db.query(models.StudentGlobalProfile).join(models.StudentProfile).join(models.User).filter(
            or_(
                models.StudentProfile.registration_number == registration,
                models.User.email == email if email else False,
            )
        ).first()
        duplicate_count += 1 if existing else 0
        preview.append({
            "row": index,
            "full_name": full_name,
            "registration_number": registration,
            "email": email or None,
            "date_of_birth": row.get("date_of_birth") or row.get("date_naissance"),
            "gender": row.get("gender") or row.get("genre"),
            "class_id": row.get("class_id"),
            "existing_global_profile_id": existing.id if existing else None,
        })
    batch = models.StudentImportBatch(
        school_id=context.school_id,
        school_model_assignment_id=context.school_model_assignment_id,
        academic_year_id=context.academic_year_id,
        filename=file.filename or "import",
        source_format=(file.filename or "").rsplit(".", 1)[-1].lower(),
        preview_payload=preview,
        error_payload=errors,
        duplicate_count=duplicate_count,
        created_by_user_id=current_user.id,
    )
    db.add(batch)
    db.flush()
    audit.record_audit(
        db,
        action="student_import.previewed",
        current_user=current_user,
        entity_type="student_import_batch",
        entity_id=batch.id,
        details={"rows": len(preview), "errors": len(errors), "duplicates": duplicate_count},
    )
    db.commit()
    return {"batch_id": batch.id, "rows": preview, "errors": errors, "duplicate_count": duplicate_count}


@router.post("/imports/commit")
def commit_student_import(
    payload: schemas.StudentImportCommit,
    current_user: models.User = Depends(security.get_current_user),
    db: Session = Depends(database.get_db),
):
    rbac.require_permission(current_user, "students:import", db)
    if not payload.confirm:
        raise HTTPException(status_code=400, detail="Confirmation obligatoire.")
    batch = db.query(models.StudentImportBatch).filter(
        models.StudentImportBatch.id == payload.batch_id,
        models.StudentImportBatch.status == "preview",
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Lot d'import introuvable ou deja traite.")
    context = school_context.resolve_context(
        db,
        current_user,
        school_model_assignment_id=batch.school_model_assignment_id,
        academic_year_id=batch.academic_year_id,
    )
    imported = 0
    linked = 0
    try:
        for row in batch.preview_payload:
            global_profile = None
            if row.get("existing_global_profile_id"):
                global_profile = db.query(models.StudentGlobalProfile).filter(
                    models.StudentGlobalProfile.id == row["existing_global_profile_id"]
                ).first()
                linked += 1
            else:
                email = row.get("email") or f"{row['registration_number'].lower()}@import.teducai.local"
                user = models.User(
                    email=email,
                    full_name=row["full_name"],
                    role=models.UserRole.STUDENT,
                    school_id=context.school_id,
                    hashed_password=security.get_password_hash(secrets.token_urlsafe(24)),
                    is_active=True,
                )
                db.add(user)
                db.flush()
                birth = None
                if row.get("date_of_birth"):
                    try:
                        birth = datetime.fromisoformat(str(row["date_of_birth"]))
                    except ValueError:
                        birth = None
                legacy = models.StudentProfile(
                    user_id=user.id,
                    school_model_assignment_id=context.school_model_assignment_id,
                    registration_number=row["registration_number"],
                    date_of_birth=birth,
                    gender=row.get("gender"),
                    parent_name="A completer",
                    parent_phone="",
                    current_class_id=int(row["class_id"]) if row.get("class_id") else None,
                )
                db.add(legacy)
                db.flush()
                global_profile = student_lifecycle.ensure_global_profile(db, legacy)
                imported += 1
            existing_enrollment = student_lifecycle.enrollment_for_context(
                db,
                global_profile.id,
                school_id=context.school_id,
                school_model_assignment_id=context.school_model_assignment_id,
                academic_year_id=context.academic_year_id,
            )
            if not existing_enrollment:
                student_lifecycle.ensure_current_enrollment(
                    db,
                    student_profile=global_profile.student_profile,
                    current_user=current_user,
                    school_id=context.school_id,
                    school_model_assignment_id=context.school_model_assignment_id,
                    academic_year_id=context.academic_year_id,
                    class_id=int(row["class_id"]) if row.get("class_id") else None,
                )
        batch.status = "completed"
        batch.imported_count = imported
        batch.completed_at = _now()
        audit.record_audit(
            db,
            action="student_import.completed",
            current_user=current_user,
            entity_type="student_import_batch",
            entity_id=batch.id,
            details={"created_profiles": imported, "linked_existing": linked},
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"batch_id": batch.id, "created_profiles": imported, "linked_existing": linked, "status": "completed"}
