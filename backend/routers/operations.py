from datetime import datetime
import csv
from io import BytesIO, StringIO
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from .. import audit, database, models, schemas, security

router = APIRouter(prefix="/operations", tags=["Institution Operations"])

OPERATIONS_ROLES = {
    models.UserRole.SUPER_ADMIN,
    models.UserRole.SCHOOL_ADMIN,
    models.UserRole.DIRECTION,
    models.UserRole.REGISTRAR,
}


def _school_id(current_user: models.User) -> int:
    if not current_user.school_id:
        raise HTTPException(status_code=400, detail="School context is required")
    return current_user.school_id


def _ensure_manager(current_user: models.User) -> None:
    if current_user.role not in OPERATIONS_ROLES:
        raise HTTPException(status_code=403, detail="Not authorized")


def _rows_from_upload(upload: UploadFile) -> list[dict]:
    content = upload.file.read()
    filename = upload.filename or ""
    if filename.lower().endswith(".xlsx"):
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise HTTPException(status_code=400, detail="XLSX import requires openpyxl") from exc
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    return list(csv.DictReader(StringIO(content.decode("utf-8-sig"))))


def _inventory_status(quantity: int, minimum_quantity: int) -> models.InventoryStatus:
    if quantity <= 0:
        return models.InventoryStatus.OUT_OF_STOCK
    if quantity <= minimum_quantity:
        return models.InventoryStatus.LOW_STOCK
    return models.InventoryStatus.AVAILABLE


@router.get("/programs", response_model=List[schemas.AcademicProgramResponse])
def list_programs(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    return db.query(models.AcademicProgram).filter(models.AcademicProgram.school_id == _school_id(current_user)).order_by(models.AcademicProgram.name.asc()).all()


@router.post("/programs", response_model=schemas.AcademicProgramResponse)
def create_program(program: schemas.AcademicProgramCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.AcademicProgram(**program.model_dump(), school_id=_school_id(current_user))
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/admissions", response_model=List[schemas.AdmissionApplicationResponse])
def list_admissions(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    return db.query(models.AdmissionApplication).filter(models.AdmissionApplication.school_id == _school_id(current_user)).order_by(models.AdmissionApplication.created_at.desc()).all()


@router.post("/admissions", response_model=schemas.AdmissionApplicationResponse)
def create_admission(application: schemas.AdmissionApplicationCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.AdmissionApplication(**application.model_dump(), school_id=_school_id(current_user), handled_by_id=current_user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/admissions/{application_id}", response_model=schemas.AdmissionApplicationResponse)
def update_admission(application_id: int, update: schemas.AdmissionApplicationUpdate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = db.query(models.AdmissionApplication).filter(models.AdmissionApplication.id == application_id, models.AdmissionApplication.school_id == _school_id(current_user)).first()
    if not row:
        raise HTTPException(status_code=404, detail="Admission application not found")
    row.status = update.status
    row.notes = update.notes
    row.handled_by_id = current_user.id
    db.commit()
    db.refresh(row)
    return row


@router.post("/admissions/{application_id}/enroll", response_model=schemas.AdmissionEnrollmentResponse)
def enroll_admission(application_id: int, payload: schemas.AdmissionEnrollmentCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    school_id = _school_id(current_user)
    application = db.query(models.AdmissionApplication).filter(
        models.AdmissionApplication.id == application_id,
        models.AdmissionApplication.school_id == school_id,
    ).first()
    if not application:
        raise HTTPException(status_code=404, detail="Admission application not found")
    if application.status not in [models.AdmissionStatus.SUBMITTED, models.AdmissionStatus.ACCEPTED]:
        raise HTTPException(status_code=400, detail="Only submitted or accepted applications can be enrolled")
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(status_code=400, detail="Email already exists")
    security.validate_password_strength(payload.password)

    cls = None
    if payload.class_id:
        cls = db.query(models.Class).filter(models.Class.id == payload.class_id, models.Class.school_id == school_id).first()
        if not cls:
            raise HTTPException(status_code=404, detail="Class not found")

    student_user = models.User(
        email=payload.email,
        hashed_password=security.get_password_hash(payload.password),
        full_name=payload.full_name or application.applicant_name,
        role=models.UserRole.STUDENT,
        school_id=school_id,
        phone_number=application.applicant_phone,
    )
    db.add(student_user)
    db.flush()

    student = models.StudentProfile(
        user_id=student_user.id,
        registration_number=payload.registration_number or f"ADM-{application.id:05d}-{student_user.id:05d}",
        date_of_birth=payload.date_of_birth,
        gender=payload.gender,
        parent_name=application.applicant_name,
        parent_phone=application.applicant_phone,
        parent_email=application.applicant_email,
        previous_level=application.desired_level,
        current_class_id=cls.id if cls else None,
        status=models.StudentStatus.ASSIGNED if cls else models.StudentStatus.UNASSIGNED,
    )
    db.add(student)
    db.flush()

    documents_count = 0
    if payload.create_registration_documents:
        for name in ["Extrait de naissance", "Bulletin scolaire precedent", "Piece d'identite parent ou representant"]:
            db.add(models.StudentRegistrationDocument(student_id=student.id, name=name, is_received=False, updated_by_id=current_user.id))
            documents_count += 1

    generated_fees = 0
    if payload.generate_fees:
        schedule_query = db.query(models.FeeSchedule).filter(models.FeeSchedule.school_id == school_id, models.FeeSchedule.is_current == True)
        schedules = schedule_query.all()
        for schedule in schedules:
            if schedule.class_id and (not cls or schedule.class_id != cls.id):
                continue
            if schedule.level and (not cls or schedule.level != cls.level):
                continue
            db.add(models.Fee(
                title=schedule.name,
                amount=schedule.amount,
                category=schedule.name,
                category_order=schedule.category_order,
                is_required=schedule.is_required,
                academic_year_id=schedule.academic_year_id,
                class_id=cls.id if cls else None,
                student_id=student.id,
                school_id=school_id,
                due_date=datetime.utcnow(),
            ))
            generated_fees += 1

    application.status = models.AdmissionStatus.ENROLLED
    application.handled_by_id = current_user.id
    application.notes = (application.notes or "") + f"\nEnrolled as student #{student.id}."
    audit.record_audit(db, action="admission.enrolled", current_user=current_user, entity_type="admission_application", entity_id=application.id, details={"student_profile_id": student.id})
    db.commit()
    db.refresh(student)
    return {
        "application_id": application.id,
        "student_user_id": student_user.id,
        "student_profile_id": student.id,
        "class_id": student.current_class_id,
        "generated_fees": generated_fees,
        "registration_documents": documents_count,
    }


@router.post("/imports/students")
def import_students(
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(security.get_current_user),
):
    _ensure_manager(current_user)
    school_id = _school_id(current_user)
    rows = _rows_from_upload(file)
    created = 0
    skipped: list[dict] = []
    for index, row in enumerate(rows, start=2):
        email = str(row.get("email") or "").strip()
        full_name = str(row.get("full_name") or row.get("name") or "").strip()
        registration_number = str(row.get("registration_number") or row.get("matricule") or "").strip()
        if not email or not full_name or not registration_number:
            skipped.append({"row": index, "reason": "email, full_name and registration_number are required"})
            continue
        if db.query(models.User).filter(models.User.email == email).first():
            skipped.append({"row": index, "reason": "email already exists"})
            continue
        if db.query(models.StudentProfile).filter(models.StudentProfile.registration_number == registration_number).first():
            skipped.append({"row": index, "reason": "registration number already exists"})
            continue
        user = models.User(
            email=email,
            full_name=full_name,
            hashed_password=security.get_password_hash(str(row.get("password") or "ChangeMe123!Secure")),
            role=models.UserRole.STUDENT,
            school_id=school_id,
        )
        db.add(user)
        db.flush()
        db.add(models.StudentProfile(
            user_id=user.id,
            registration_number=registration_number,
            parent_name=str(row.get("parent_name") or "Parent"),
            parent_phone=str(row.get("parent_phone") or "+2250102030405"),
            gender=str(row.get("gender") or "N/A"),
            status=models.StudentStatus.UNASSIGNED,
        ))
        created += 1
    db.commit()
    return {"created": created, "skipped": skipped}


@router.get("/exams", response_model=List[schemas.ExamSessionResponse])
def list_exam_sessions(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    return db.query(models.ExamSession).filter(models.ExamSession.school_id == _school_id(current_user)).order_by(models.ExamSession.created_at.desc()).all()


@router.post("/exams", response_model=schemas.ExamSessionResponse)
def create_exam_session(exam: schemas.ExamSessionCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.ExamSession(**exam.model_dump(), school_id=_school_id(current_user), created_by_id=current_user.id)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/inventory", response_model=List[schemas.InventoryItemResponse])
def list_inventory(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    return db.query(models.InventoryItem).filter(models.InventoryItem.school_id == _school_id(current_user)).order_by(models.InventoryItem.category.asc(), models.InventoryItem.name.asc()).all()


@router.post("/inventory", response_model=schemas.InventoryItemResponse)
def create_inventory_item(item: schemas.InventoryItemCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.InventoryItem(
        **item.model_dump(),
        status=_inventory_status(item.quantity, item.minimum_quantity),
        school_id=_school_id(current_user),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/payroll", response_model=List[schemas.PayrollRecordResponse])
def list_payroll(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    return db.query(models.PayrollRecord).filter(models.PayrollRecord.school_id == _school_id(current_user)).order_by(models.PayrollRecord.created_at.desc()).all()


@router.post("/payroll", response_model=schemas.PayrollRecordResponse)
def create_payroll(record: schemas.PayrollRecordCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    staff = db.query(models.User).filter(models.User.id == record.staff_user_id, models.User.school_id == _school_id(current_user)).first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff user not found")
    row = models.PayrollRecord(
        **record.model_dump(),
        net_amount=record.gross_amount - record.deductions,
        school_id=_school_id(current_user),
        created_by_id=current_user.id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.put("/payroll/{record_id}", response_model=schemas.PayrollRecordResponse)
def update_payroll(record_id: int, update: schemas.PayrollRecordUpdate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = db.query(models.PayrollRecord).filter(models.PayrollRecord.id == record_id, models.PayrollRecord.school_id == _school_id(current_user)).first()
    if not row:
        raise HTTPException(status_code=404, detail="Payroll record not found")
    row.status = update.status
    row.paid_at = datetime.utcnow() if update.status == models.PayrollStatus.PAID else row.paid_at
    db.commit()
    db.refresh(row)
    return row


@router.get("/transport", response_model=List[schemas.TransportRouteResponse])
def list_transport(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    return db.query(models.TransportRoute).filter(models.TransportRoute.school_id == _school_id(current_user)).order_by(models.TransportRoute.name.asc()).all()


@router.post("/transport", response_model=schemas.TransportRouteResponse)
def create_transport(route: schemas.TransportRouteCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.TransportRoute(**route.model_dump(), school_id=_school_id(current_user))
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/canteen", response_model=List[schemas.CanteenMealPlanResponse])
def list_canteen(db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    return db.query(models.CanteenMealPlan).filter(models.CanteenMealPlan.school_id == _school_id(current_user)).order_by(models.CanteenMealPlan.day_of_week.asc().nullslast(), models.CanteenMealPlan.name.asc()).all()


@router.post("/canteen", response_model=schemas.CanteenMealPlanResponse)
def create_canteen_plan(plan: schemas.CanteenMealPlanCreate, db: Session = Depends(database.get_db), current_user: models.User = Depends(security.get_current_user)):
    _ensure_manager(current_user)
    row = models.CanteenMealPlan(**plan.model_dump(), school_id=_school_id(current_user))
    db.add(row)
    db.commit()
    db.refresh(row)
    return row
